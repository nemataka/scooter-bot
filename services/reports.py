import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from database.db import get_all_orders, get_user
from utils.helpers import format_date, format_amount, get_status_text


async def generate_excel_report():
    """Excel отчёт яратиш"""
    
    os.makedirs("reports", exist_ok=True)
    
    filename = f"orders_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join("reports", filename)
    
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Барча заказлар"
    
    headers = ['ID', 'Мижоз', 'Телефон', 'Манзил', 'Буюртма', 
               'Сумма', 'Статус', 'Яратилган', 'Бажарилган', 'Курьер']
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    orders = await get_all_orders()
    for row, order in enumerate(orders, 2):
        order_id = order[0]
        client_name = order[1]
        client_phone = order[2]
        address = order[3]
        details = order[4]
        amount = order[5] or 0
        status = get_status_text(order[6])
        created_at = format_date(order[9])
        completed_at = format_date(order[11]) if len(order) > 11 else "—"
        
        courier_name = "—"
        courier_id = order[7]
        if courier_id:
            courier = await get_user(courier_id)
            if courier:
                courier_name = courier[2] or courier[1] or f"ID:{courier_id}"
        
        data = [order_id, client_name, client_phone, address, details,
                amount, status, created_at, completed_at, courier_name]
        
        for col, value in enumerate(data, 1):
            cell = ws1.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
    
    widths = [5, 20, 15, 25, 30, 12, 15, 18, 18, 20]
    for col, width in enumerate(widths, 1):
        ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = width
    
    wb.save(filepath)
    return filepath